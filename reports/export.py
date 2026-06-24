from io import BytesIO

from openpyxl import Workbook

from feedback.export_utils import (
    _auto_width,
    _write_feedback_attachments_sheet,
    _write_form_feedback_sheet,
    _write_response_feedback_sheet,
    build_form_feedback_workbook,
    build_response_feedback_workbook,
)
from feedback.query_utils import get_feedback_list_queryset, get_response_feedback_list_queryset

from .queries import (
    query_admin_audit_log,
    query_chat_activity,
    query_citation_accuracy,
    query_document_uploads,
    query_failed_uploads,
    query_kb_usage,
    query_language_issues,
    query_login_activity,
    query_negative_feedback,
    query_user_activity,
    query_user_list,
)


def _write_rows(ws, headers, rows, row_builder):
    ws.append(headers)
    for item in rows:
        ws.append(row_builder(item))
    _auto_width(ws)


def _append_sheet(workbook, title, headers, rows, row_builder):
    ws = workbook.create_sheet(title[:31])
    _write_rows(ws, headers, rows, row_builder)
    return ws


def build_single_report_workbook(slug, params, request=None):
    if slug == "form_feedback":
        return build_form_feedback_workbook(params, request=request)
    if slug == "response_feedback":
        return build_response_feedback_workbook(params, request=request)

    workbook = Workbook()
    ws = workbook.active

    if slug == "chat_activity":
        rows, _ = query_chat_activity(params)
        ws.title = "Chat Activity"
        _write_rows(
            ws,
            ["Date", "Chat Count", "Active Users"],
            rows,
            lambda r: [r["date"], r["chat_count"], r["active_users"]],
        )
    elif slug == "kb_usage":
        rows, _ = query_kb_usage(params)
        ws.title = "KB Usage"
        _write_rows(
            ws,
            ["ID", "Name", "Chats", "Files", "Completed", "Failed", "Created At"],
            rows,
            lambda r: [
                r["id"],
                r["name"],
                r["chat_count"],
                r["file_count"],
                r["completed_files"],
                r["failed_files"],
                r["created_at"],
            ],
        )
    elif slug == "user_activity":
        rows, _ = query_user_activity(params)
        ws.title = "User Activity"
        _write_rows(
            ws,
            ["User ID", "Email", "Username", "Chat Count"],
            rows,
            lambda r: [r["user_id"], r["email"], r["username"], r["chat_count"]],
        )
    elif slug == "document_uploads":
        rows, _ = query_document_uploads(params)
        ws.title = "Document Uploads"
        _write_rows(
            ws,
            ["ID", "File Name", "Status", "Knowledge Base", "Chat ID", "Uploaded At"],
            rows,
            lambda r: [
                r["id"],
                r["file_name"],
                r["status"],
                r["knowledge_base"],
                r["chat_id"],
                r["uploaded_at"],
            ],
        )
    elif slug == "failed_uploads":
        rows, _ = query_failed_uploads(params)
        ws.title = "Failed Uploads"
        _write_rows(
            ws,
            ["ID", "File Name", "Knowledge Base", "Error", "Technical Error", "Uploaded At"],
            rows,
            lambda r: [
                r["id"],
                r["file_name"],
                r["knowledge_base"],
                r["error"],
                r["technical_error"],
                r["uploaded_at"],
            ],
        )
    elif slug == "user_list":
        rows, _ = query_user_list(params)
        ws.title = "Users"
        _write_rows(
            ws,
            [
                "ID",
                "Username",
                "Email",
                "First Name",
                "Last Name",
                "Active",
                "Staff",
                "Superuser",
                "Last Login",
                "Date Joined",
            ],
            rows,
            lambda r: [
                r["id"],
                r["username"],
                r["email"],
                r["first_name"],
                r["last_name"],
                r["is_active"],
                r["is_staff"],
                r["is_superuser"],
                r["last_login"],
                r["date_joined"],
            ],
        )
    elif slug == "admin_audit_log":
        rows, _ = query_admin_audit_log(params)
        ws.title = "Admin Audit Log"
        _write_rows(
            ws,
            ["ID", "Action Time", "User", "Action", "Object", "Change Message"],
            rows,
            lambda r: [
                r["id"],
                r["action_time"],
                r["user"],
                r["action"],
                r["object_repr"],
                r["change_message"],
            ],
        )
    elif slug == "login_activity":
        rows, _ = query_login_activity(params)
        ws.title = "Login Activity"
        _write_rows(
            ws,
            ["ID", "Username", "Email", "Active", "Staff", "Last Login", "Days Since Login", "Date Joined"],
            rows,
            lambda r: [
                r["id"],
                r["username"],
                r["email"],
                r["is_active"],
                r["is_staff"],
                r["last_login"],
                r["days_since_login"],
                r["date_joined"],
            ],
        )
    elif slug == "negative_feedback":
        data, _ = query_negative_feedback(params)
        ws.title = "Negative Breakdown"
        _write_rows(
            ws,
            ["Category", "Label", "Count"],
            data["breakdown"],
            lambda r: [r["category"], r["category_label"], r["count"]],
        )
        _append_sheet(
            workbook,
            "Negative Details",
            ["ID", "User Email", "Chat ID", "Category", "Details", "Question", "Created At"],
            data["details"],
            lambda r: [
                r["id"],
                r["user_email"],
                r["chat_id"],
                r["category_label"],
                r["details"],
                r["user_question"],
                r["created_at"],
            ],
        )
    elif slug == "citation_accuracy":
        rows, _ = query_citation_accuracy(params)
        ws.title = "Citation Accuracy"
        _write_rows(
            ws,
            ["ID", "User Email", "Chat ID", "Details", "Question", "Response", "Created At"],
            rows,
            lambda r: [
                r["id"],
                r["user_email"],
                r["chat_id"],
                r["details"],
                r["user_question"],
                r["assistant_response"],
                r["created_at"],
            ],
        )
    elif slug == "language_issues":
        rows, _ = query_language_issues(params)
        ws.title = "Language Issues"
        _write_rows(
            ws,
            ["ID", "User Email", "Chat ID", "Details", "Question", "Response", "Created At"],
            rows,
            lambda r: [
                r["id"],
                r["user_email"],
                r["chat_id"],
                r["details"],
                r["user_question"],
                r["assistant_response"],
                r["created_at"],
            ],
        )
    else:
        raise ValueError(f"Unknown report slug: {slug}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_all_reports_workbook(params, request=None):
    """Multi-sheet export with detailed row data per report (no summary-only sheet)."""
    workbook = Workbook()
    index_ws = workbook.active
    index_ws.title = "Index"
    index_ws.append(["Report", "Excel Tab", "Data Rows", "Description"])

    def _index_row(report_name, tab_name, row_count, description):
        index_ws.append([report_name, tab_name, row_count, description])

    form_qs = get_feedback_list_queryset(params)
    form_ws = workbook.create_sheet("Form Feedback")
    _write_form_feedback_sheet(form_ws, form_qs, request)
    _index_row(
        "Form Feedback Report",
        "Form Feedback",
        form_qs.count(),
        "Sidebar feedback submissions with full message details",
    )

    attachments_ws = workbook.create_sheet("Form Attachments")
    _write_feedback_attachments_sheet(attachments_ws, form_qs, request)
    attachment_count = sum(len(list(feedback.attachments.all())) for feedback in form_qs)
    _index_row(
        "Form Feedback Attachments",
        "Form Attachments",
        attachment_count,
        "One row per uploaded screenshot — click Download URL to open file",
    )

    response_qs = get_response_feedback_list_queryset(params)
    response_ws = workbook.create_sheet("Response Feedback")
    _write_response_feedback_sheet(response_ws, response_qs)
    _index_row(
        "Response Feedback Report",
        "Response Feedback",
        response_qs.count(),
        "Chat thumbs up/down with question, response, and reason",
    )

    chat_rows, chat_count = query_chat_activity(params)
    _append_sheet(
        workbook,
        "Chat Activity",
        ["Date", "Chat Count", "Active Users"],
        chat_rows,
        lambda r: [r["date"], r["chat_count"], r["active_users"]],
    )
    _index_row(
        "Chat Activity Report",
        "Chat Activity",
        chat_count,
        "Daily chat volume and active users",
    )

    kb_rows, kb_count = query_kb_usage(params)
    _append_sheet(
        workbook,
        "KB Usage",
        ["ID", "Name", "Chats", "Files", "Completed", "Failed", "Created At"],
        kb_rows,
        lambda r: [
            r["id"],
            r["name"],
            r["chat_count"],
            r["file_count"],
            r["completed_files"],
            r["failed_files"],
            r["created_at"],
        ],
    )
    _index_row(
        "Knowledge Base Usage Report",
        "KB Usage",
        kb_count,
        "Knowledge base usage by chats and files",
    )

    user_activity_rows, user_activity_count = query_user_activity(params)
    _append_sheet(
        workbook,
        "User Activity",
        ["User ID", "Email", "Username", "Chat Count"],
        user_activity_rows,
        lambda r: [r["user_id"], r["email"], r["username"], r["chat_count"]],
    )
    _index_row(
        "User Activity Report",
        "User Activity",
        user_activity_count,
        "Chats per user in selected period",
    )

    upload_rows, upload_count = query_document_uploads(params)
    _append_sheet(
        workbook,
        "Document Uploads",
        ["ID", "File Name", "Status", "Knowledge Base", "Chat ID", "Uploaded At"],
        upload_rows,
        lambda r: [
            r["id"],
            r["file_name"],
            r["status"],
            r["knowledge_base"],
            r["chat_id"],
            r["uploaded_at"],
        ],
    )
    _index_row(
        "Document Upload Report",
        "Document Uploads",
        upload_count,
        "All uploaded documents with status",
    )

    failed_rows, failed_count = query_failed_uploads(params)
    _append_sheet(
        workbook,
        "Failed Uploads",
        ["ID", "File Name", "Knowledge Base", "Error", "Technical Error", "Uploaded At"],
        failed_rows,
        lambda r: [
            r["id"],
            r["file_name"],
            r["knowledge_base"],
            r["error"],
            r["technical_error"],
            r["uploaded_at"],
        ],
    )
    _index_row(
        "Failed Upload Report",
        "Failed Uploads",
        failed_count,
        "Failed PDF uploads with error details",
    )

    user_rows, user_count = query_user_list(params)
    _append_sheet(
        workbook,
        "Users",
        [
            "ID",
            "Username",
            "Email",
            "First Name",
            "Last Name",
            "Active",
            "Staff",
            "Superuser",
            "Last Login",
            "Date Joined",
        ],
        user_rows,
        lambda r: [
            r["id"],
            r["username"],
            r["email"],
            r["first_name"],
            r["last_name"],
            r["is_active"],
            r["is_staff"],
            r["is_superuser"],
            r["last_login"],
            r["date_joined"],
        ],
    )
    _index_row("User List Export", "Users", user_count, "All registered users with roles")

    audit_rows, audit_count = query_admin_audit_log(params)
    _append_sheet(
        workbook,
        "Admin Audit Log",
        ["ID", "Action Time", "User", "Action", "Object", "Change Message"],
        audit_rows,
        lambda r: [
            r["id"],
            r["action_time"],
            r["user"],
            r["action"],
            r["object_repr"],
            r["change_message"],
        ],
    )
    _index_row("Admin Audit Log", "Admin Audit Log", audit_count, "Django admin actions")

    login_rows, login_count = query_login_activity(params)
    _append_sheet(
        workbook,
        "Login Activity",
        ["ID", "Username", "Email", "Active", "Staff", "Last Login", "Days Since Login", "Date Joined"],
        login_rows,
        lambda r: [
            r["id"],
            r["username"],
            r["email"],
            r["is_active"],
            r["is_staff"],
            r["last_login"],
            r["days_since_login"],
            r["date_joined"],
        ],
    )
    _index_row("Login Activity Report", "Login Activity", login_count, "User last login and account status")

    negative_data, negative_count = query_negative_feedback(params)
    _append_sheet(
        workbook,
        "Negative Breakdown",
        ["Category", "Label", "Count"],
        negative_data["breakdown"],
        lambda r: [r["category"], r["category_label"], r["count"]],
    )
    _append_sheet(
        workbook,
        "Negative Details",
        ["ID", "User Email", "Chat ID", "Category", "Details", "Question", "Created At"],
        negative_data["details"],
        lambda r: [
            r["id"],
            r["user_email"],
            r["chat_id"],
            r["category_label"],
            r["details"],
            r["user_question"],
            r["created_at"],
        ],
    )
    _index_row(
        "Negative Feedback Analysis",
        "Negative Breakdown + Negative Details",
        negative_count,
        "Thumbs down breakdown and complaint details",
    )

    citation_rows, citation_count = query_citation_accuracy(params)
    _append_sheet(
        workbook,
        "Citation Accuracy",
        ["ID", "User Email", "Chat ID", "Details", "Question", "Response", "Created At"],
        citation_rows,
        lambda r: [
            r["id"],
            r["user_email"],
            r["chat_id"],
            r["details"],
            r["user_question"],
            r["assistant_response"],
            r["created_at"],
        ],
    )
    _index_row(
        "Citation Accuracy Report",
        "Citation Accuracy",
        citation_count,
        "Wrong document/source complaints",
    )

    language_rows, language_count = query_language_issues(params)
    _append_sheet(
        workbook,
        "Language Issues",
        ["ID", "User Email", "Chat ID", "Details", "Question", "Response", "Created At"],
        language_rows,
        lambda r: [
            r["id"],
            r["user_email"],
            r["chat_id"],
            r["details"],
            r["user_question"],
            r["assistant_response"],
            r["created_at"],
        ],
    )
    _index_row(
        "Language Issue Report",
        "Language Issues",
        language_count,
        "Language and translation feedback",
    )

    _auto_width(index_ws)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
