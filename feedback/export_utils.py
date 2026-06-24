from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .constants import FEEDBACK_CATEGORIES, RESPONSE_FEEDBACK_CATEGORIES
from .query_utils import get_feedback_list_queryset, get_response_feedback_list_queryset
from .utils import build_feedback_attachment_download_url

FORM_FEEDBACK_HEADERS = [
    "ID",
    "Name",
    "Email",
    "Logged-in User Email",
    "Category",
    "Subject",
    "Message",
    "Page URL",
    "Status",
    "Attachment Count",
    "Attachment Filenames",
    "Attachment Download URLs",
    "Created At",
    "Updated At",
]

FEEDBACK_ATTACHMENT_HEADERS = [
    "Feedback ID",
    "Subject",
    "Submitter Email",
    "Filename",
    "Download URL",
    "File Size (bytes)",
    "Uploaded At",
]

RESPONSE_FEEDBACK_HEADERS = [
    "ID",
    "User Email",
    "Username",
    "Chat ID",
    "Message ID",
    "Rating",
    "Category",
    "Details",
    "User Question",
    "Assistant Response",
    "Created At",
    "Updated At",
]

SUMMARY_HEADERS = ["Metric", "Value"]

_LINK_FONT = Font(color="0563C1", underline="single")


def _auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(length + 2, 12), 60
        )


def _category_label(value, choices):
    return dict(choices).get(value, value)


def _set_hyperlink_cell(ws, row, column, url):
    if not url:
        return
    cell = ws.cell(row=row, column=column, value=url)
    cell.hyperlink = url
    cell.font = _LINK_FONT


def _attachment_urls(feedback, request=None):
    urls = []
    names = []
    for attachment in feedback.attachments.all():
        names.append(attachment.original_filename)
        urls.append(
            build_feedback_attachment_download_url(
                attachment, request=request, signed=True
            )
        )
    return names, urls


def _write_form_feedback_sheet(ws, queryset, request=None):
    ws.append(FORM_FEEDBACK_HEADERS)
    url_col = FORM_FEEDBACK_HEADERS.index("Attachment Download URLs") + 1
    for row_idx, feedback in enumerate(queryset, start=2):
        attachment_names, attachment_urls = _attachment_urls(feedback, request)
        ws.append(
            [
                str(feedback.id),
                feedback.name,
                feedback.email,
                feedback.user.email if feedback.user_id else "",
                _category_label(feedback.category, FEEDBACK_CATEGORIES),
                feedback.subject,
                feedback.message,
                feedback.page_url,
                feedback.status,
                len(attachment_names),
                " | ".join(attachment_names),
                " | ".join(attachment_urls),
                feedback.created_at.isoformat(),
                feedback.updated_at.isoformat(),
            ]
        )
        if attachment_urls:
            _set_hyperlink_cell(ws, row_idx, url_col, attachment_urls[0])
    _auto_width(ws)


def _write_feedback_attachments_sheet(ws, queryset, request=None):
    ws.append(FEEDBACK_ATTACHMENT_HEADERS)
    url_col = FEEDBACK_ATTACHMENT_HEADERS.index("Download URL") + 1
    row_idx = 2
    for feedback in queryset:
        for attachment in feedback.attachments.all():
            url = build_feedback_attachment_download_url(
                attachment, request=request, signed=True
            )
            ws.append(
                [
                    str(feedback.id),
                    feedback.subject,
                    feedback.email,
                    attachment.original_filename,
                    url,
                    attachment.file_size,
                    attachment.created_at.isoformat(),
                ]
            )
            _set_hyperlink_cell(ws, row_idx, url_col, url)
            row_idx += 1
    _auto_width(ws)


def _write_response_feedback_sheet(ws, queryset):
    ws.append(RESPONSE_FEEDBACK_HEADERS)
    for item in queryset:
        ws.append(
            [
                str(item.id),
                item.user.email if item.user_id else "",
                item.user.username if item.user_id else "",
                item.chat_id,
                item.message_id,
                item.rating,
                _category_label(item.category, RESPONSE_FEEDBACK_CATEGORIES),
                item.details,
                item.user_question,
                item.assistant_response,
                item.created_at.isoformat(),
                item.updated_at.isoformat(),
            ]
        )
    _auto_width(ws)


def _write_summary_sheet(ws, form_qs, response_qs):
    form_count = form_qs.count()
    response_count = response_qs.count()
    thumbs_up = response_qs.filter(rating="up").count()
    thumbs_down = response_qs.filter(rating="down").count()

    ws.append(SUMMARY_HEADERS)
    rows = [
        ("Form feedback submissions", form_count),
        ("Response feedback (thumbs) submissions", response_count),
        ("Helpful (thumbs up)", thumbs_up),
        ("Not helpful (thumbs down)", thumbs_down),
        ("Form feedback — new", form_qs.filter(status="new").count()),
        ("Form feedback — in review", form_qs.filter(status="in_review").count()),
        ("Form feedback — resolved", form_qs.filter(status="resolved").count()),
    ]
    for label, value in rows:
        ws.append([label, value])
    _auto_width(ws)


def build_feedback_workbook(params, request=None):
    form_qs = get_feedback_list_queryset(params)
    response_qs = get_response_feedback_list_queryset(params)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    _write_summary_sheet(summary_ws, form_qs, response_qs)

    form_ws = workbook.create_sheet("Form Feedback")
    _write_form_feedback_sheet(form_ws, form_qs, request)

    attachments_ws = workbook.create_sheet("Form Attachments")
    _write_feedback_attachments_sheet(attachments_ws, form_qs, request)

    response_ws = workbook.create_sheet("Response Feedback")
    _write_response_feedback_sheet(response_ws, response_qs)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer, form_qs.count(), response_qs.count()


def build_form_feedback_workbook(params, request=None):
    form_qs = get_feedback_list_queryset(params)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Form Feedback"
    _write_form_feedback_sheet(ws, form_qs, request)

    attachments_ws = workbook.create_sheet("Form Attachments")
    _write_feedback_attachments_sheet(attachments_ws, form_qs, request)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_response_feedback_workbook(params, request=None):
    response_qs = get_response_feedback_list_queryset(params)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Response Feedback"
    _write_response_feedback_sheet(ws, response_qs)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
